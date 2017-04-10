#This file will contain all initial work done to connect to Excel Sheet
import openpyxl

#Returns a pointer to Sheet we're concerned with
def file_connect():
	workbook = openpyxl.load_workbook('C:/Users/Nick/Documents/ATO_Docs/ATO_Alumni_Full_Contact_List_Stewart_Howe_Elevate.xlsx')
	sheetnames = workbook.get_sheet_names()
	ATOSheet = workbook.get_sheet_by_name('ATOME123')
	print (sheetnames)
	print (ATOSheet.title)
	return ATOSheet