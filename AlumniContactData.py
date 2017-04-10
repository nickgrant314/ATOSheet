import sys
import codecs
import openpyxl


workbook = openpyxl.load_workbook('C:/Users/Nick/Documents/ATO_Docs/ATO_Alumni_Full_Contact_List_Stewart_Howe_Elevate.xlsx')
sheetnames = workbook.get_sheet_names()
ATOSheet = workbook.get_sheet_by_name('ATOME123')
print (sheetnames)
print (ATOSheet.title)

#Put in a function due to reuse
def help_text():
	print("If you want Alumni info based on state, type 'State Lookup'")
	print("If you want to retrieve a list of Alumni without current phone info, type 'Invalid Phone Lookup'")
	print("If you would like to exit this experience, type 'I\'m done'")
	print()
	print("If you ever need help remembering function names, type 'Help'")

	

def user_prompt():
	return input("So what would you like to do? ")

#User inputs the initials of a state, function ouputs contact info of Alumni who currently reside there
def info_by_state(state_initials):

	stateCount = 0

	for row in range(2, ATOSheet.max_row + 1):
		state = ATOSheet['J' + str(row)].value

		if(state == state_initials):

			FirstName = ATOSheet['A' + str(row)].value
			LastName = ATOSheet['C' + str(row)].value
			MiddleInitial = ATOSheet['B' + str(row)].value
			if(MiddleInitial == None):
				MiddleInitial = ""		#Can't print "NoneTypes", if nothing found print empty string

			Phone = ATOSheet['O' + str(row)].value
			if(Phone == None):
				Phone = "#No Phone Info Available#"		#Can't print "NoneTypes", if nothing found print empty string

			#At some point replace this with an output to text file, or Excel spreadsheet?
			print (FirstName + " " + MiddleInitial + " " + LastName + "  " + Phone)
			stateCount += 1

	return stateCount


#It is assumed that the input to this function is 1 Excel sheet, with a phone numbers column
def invalid_phone_numbers(sheet):
	invalid_count = 0
	for row in range(2, ATOSheet.max_row + 1):
		phone = ATOSheet['O' + str(row)].value
		if(phone == None):
			FirstName = ATOSheet['A' + str(row)].value
			LastName = ATOSheet['C' + str(row)].value
			MiddleInitial = ATOSheet['B' + str(row)].value
			if(MiddleInitial == None):
				MiddleInitial = ""		#Can't print "NoneTypes", if nothing found print empty string

			#At some point replace this with an output to text file, or Excel spreadsheet?
			print (FirstName + " " + MiddleInitial + " " + LastName)
			invalid_count += 1

	print ("Number of people without phone number info: %d" %invalid_count)


#Main Program begins here
help_text()
userChoice = user_prompt()

while(userChoice != "I'm done"):
	if(userChoice == "State Lookup"):
		state = input("What state do you want ATO Alumni about? ")
		tally = info_by_state(state)

		print("You chose the state: %s" %state)
		print("%d ATO Alumni reside there, according to our data." %tally)

	elif(userChoice == "Invalid Phone Lookup"):
		invalid_phone_numbers(ATOSheet)

	elif(userChoice == "Help"):
		help_text()

	elif(userChoice == "I'm done"):
		break

	else:
		print("You've entered an invalid command Please try again.")
		help_text()

	userChoice = user_prompt()

print("Thank you for accessing the ATO Alumni Database. Please come back again some time!")







